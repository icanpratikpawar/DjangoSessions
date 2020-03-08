from django.shortcuts import redirect, render, render_to_response
from django.http import Http404, HttpResponse, HttpResponseRedirect, request
from .models import Login, Books
from django.core.exceptions import ObjectDoesNotExist, PermissionDenied, ValidationError
from django.core.mail import send_mail
from django.conf import settings
from django.urls import resolve
from openpyxl import load_workbook

workbook = load_workbook(
    filename="/home/pratik/Workspace/Djangoprograms/library_management/static/excelfiles/book_manage.xlsx")
workbook.active = 1
sheet = workbook.active

b = ""
# Decorators function to prohibit illegal access to the urls directly


def prohibit_url_access(func):
    def wrap(request, *args, **kwargs):
        if "user" in request.session:
            return func(request, *args, **kwargs)
        else:
            raise PermissionDenied
    return wrap


# def check_url(func):
#     def wrap(request, *args, **kwargs):
#         if resolve(request.path_info).url_name == "index" and "user" in request.session:
#             return redirect('/home')
#         else:
#             print("hii")
#     return wrap


def facebook(request):
    return HttpResponseRedirect("http://facebook.com")


def twitter(request):
    return HttpResponseRedirect("http://twitter.com")


def password_reset(request, user_name):
    a = Login.objects.filter(user_name=user_name)
    if a:
        a.update(pass_word=request.POST['pswd'])
        return render(request, 'index.html')


def password_change(request, user_name):
    a = Login.objects.filter(user_name=user_name)
    if a:
        return render(request, 'password_change.html', {'user_name': user_name})
    else:
        pass


def email_verify(request):
    # Class(Object) to get the url name
    current_url = resolve(request.path_info).url_name
    if current_url == "emailverify":
        return render(request, 'emailverify.html')


def send_email(request):
    a = Login.objects.filter(email_id=request.POST['send_email'])
    if a:
        b = "http://127.0.0.1:8000/password_change/"+a[0].user_name
        send_mail("Hello "+a[0].user_name, "This is an automated message, It A password reset for your account was requested. Please click the button below to change your password.\n "+b,
                  settings.EMAIL_HOST_USER, [request.POST['send_email']])
        return render(request, 'index.html', {'msg': 'Succesfully send the reset password message'})
    else:
        return HttpResponse("!!..Not Registered..!!")


def logout(request):
    del request.session["user"]
    # return render(request, 'index.html', {'text': 'This is my text & I am its creator'})
    return redirect('/')


def index(request):
    request.session.set_expiry(0)
    if "user" in request.session:
        return redirect("/home")
    else:
        # Starting page of my website
        request.session.set_expiry(0)
        return render(request, 'index.html')


def home(request):
    # Method to verify the right login details by admin
    if "user" not in request.session:
        try:
            a = Login.objects.filter(
                user_name=request.POST['username'], pass_word=request.POST['pswd'])
            if a:
                request.session["user"] = a[0].user_name

                return render(request, "home.html", {'user': request.session["user"]})
            else:
                return render(request, 'index.html', {'User': 'Check the username and password'})
        except Exception:
            raise PermissionDenied

    else:
        request.session.set_expiry(0)
        return render(request, "home.html", {'user': request.session["user"]})


@prohibit_url_access
def books_validate(request):
    book_name = request.POST['bookname']
    author_name = request.POST['authorname']
    book_type = request.POST['booktype']
    # Code to store the data into excel file
    row_value = sheet.max_row+1
    sheet["B"+str(row_value)] = book_name
    sheet["C"+str(row_value)] = author_name
    sheet["D"+str(row_value)] = book_type
    # Books getting stored at the database
    Books(book_name=book_name, author_name=author_name,
          book_type=book_type).save()
    # Books getting stored at the excel file
    sheet["A"+str(row_value)] = Books.objects.filter(book_name=book_name)[
        len(Books.objects.filter(book_name=book_name))-1].id
    workbook.save(
        filename="/home/pratik/Workspace/Djangoprograms/library_management/static/excelfiles/book_manage.xlsx")
    print("Success")

    # return render(request, 'home.html', {'details': 'Books Details Submitted Succesfully'})
    return redirect("/home")


@prohibit_url_access
def book_view(request):
    #query =   # [0:2]
    shee=sheet.iter_rows(min_row=2,max_row=sheet.max_row)
    return render(request, 'bookview.html', {'books': Books.objects.all(),'sheet':shee})


@prohibit_url_access
def delete_record(request, id=None):
    for col in sheet.iter_cols(max_col=1, min_row=2, max_row=sheet.max_row):
        for rows in col:
            if rows.value == int(id):
                sheet.delete_rows(rows.row, 1)
                workbook.save(
                    filename="/home/pratik/Workspace/Djangoprograms/library_management/static/excelfiles/book_manage.xlsx")
                print("Success")
    Books.objects.filter(id=id).delete()
    querys = {'Delete': 'Record Deleted Succesfully!'}
    # return render(request, 'bookview.html',querys)
    return redirect('/bookview')


@prohibit_url_access
def update(request, id):

    return render(request, 'update-record.html', {'id': id, 'user': request.session["user"],
                                                  'books': Books.objects.get(id=id)})


@prohibit_url_access
def update_record(request, id=None):
    a = Books.objects.filter(id=id)
    if a:
        for col in sheet.iter_cols(max_col=1, min_row=2, max_row=sheet.max_row):
            for rows in col:
                if rows.value == int(id):
                    sheet["B"+str(rows.row)] = request.POST['bookname']
                    sheet["C"+str(rows.row)] = request.POST['authorname']
                    sheet["D"+str(rows.row)] = request.POST['booktype']
                    workbook.save(
                        filename="/home/pratik/Workspace/Djangoprograms/library_management/static/excelfiles/book_manage.xlsx")
                    print("Success")
        a.update(book_name=request.POST['bookname'],
                 author_name=request.POST['authorname'],
                 book_type=request.POST['booktype'])
    else:
        return HttpResponse("ERRORS")
    return redirect('/bookview')


# Code to store the data of mysql database into excel file
    # workbook = load_workbook(filename="/home/pratik/Workspace/Djangoprograms/library_management/static/excelfiles/book_manage.xlsx")
    # workbook.active = 1
    # sheet = workbook.active
    # for i in Books.objects.all():
    #     row_value = sheet.max_row+1
    #     sheet["A"+str(row_value)] = i.id
    #     sheet["B"+str(row_value)] = i.book_name
    #     sheet["C"+str(row_value)] = i.author_name
    #     sheet["D"+str(row_value)] = i.book_type
    #     workbook.save(
    #         filename="/home/pratik/Workspace/Djangoprograms/library_management/static/excelfiles/book_manage.xlsx")
